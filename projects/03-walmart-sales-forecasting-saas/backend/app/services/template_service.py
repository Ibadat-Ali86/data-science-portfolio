import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import io

class TemplateService:
    """Service to generate professional Excel templates for data upload."""

    @staticmethod
    def generate_template(template_type: str) -> io.BytesIO:
        """
        Generates an Excel template file in memory.
        
        Args:
            template_type: 'sales', 'inventory', 'financial', or 'generic'
            
        Returns:
            io.BytesIO: The Excel file content
        """
        wb = Workbook()
        
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # Create sheets
        TemplateService._create_data_sheet(wb, template_type)
        TemplateService._create_instructions_sheet(wb, template_type)
        TemplateService._create_examples_sheet(wb, template_type)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def _create_data_sheet(wb: Workbook, template_type: str):
        ws = wb.create_sheet("Data Input", 0)
        
        headers = TemplateService._get_headers(template_type)
        
        # Style headers
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo-600
        header_font = Font(color="FFFFFF", bold=True, size=11, name="Segoe UI")
        centered = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = centered
            cell.border = thin_border
            
            # Adjustment column width based on header length
            ws.column_dimensions[get_column_letter(col_idx)].width = len(header) + 5

            # Add comments/validation
            if "date" in header.lower():
                TemplateService._add_date_validation(ws, col_idx)
            elif "promotion" in header.lower():
                TemplateService._add_dropdown_validation(ws, col_idx, ["Yes", "No", "High", "Low"])

        # Freeze panes
        ws.freeze_panes = "A2"

    @staticmethod
    def _create_instructions_sheet(wb: Workbook, template_type: str):
        ws = wb.create_sheet("Instructions")
        ws.sheet_properties.tabColor = "10B981" # Emerald

        # Title
        title = ws.cell(row=1, column=1, value=f"{template_type.capitalize()} Data Upload Guide")
        title.font = Font(size=16, bold=True, color="1E293B")
        
        # Steps
        steps = [
            ("1. Use the 'Data Input' sheet", "Enter your data there. Do not rename columns."),
            ("2. Date Format", "Use YYYY-MM-DD (e.g., 2024-03-25)."),
            ("3. No Calculations", "Paste values only. Formulas may cause errors."),
            ("4. Required Columns", "Date and Target (Sales/Stock/Revenue) are mandatory."),
        ]
        
        row = 3
        for step, desc in steps:
            ws.cell(row=row, column=1, value=step).font = Font(bold=True)
            ws.cell(row=row, column=2, value=desc)
            row += 1
            
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 60

    @staticmethod
    def _create_examples_sheet(wb: Workbook, template_type: str):
        ws = wb.create_sheet("Examples")
        ws.sheet_properties.tabColor = "F59E0B" # Amber
        
        headers = TemplateService._get_headers(template_type)
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header).font = Font(bold=True)
            
        # Sample data (10 rows as requested)
        samples = TemplateService._get_sample_data(template_type)
        for row_idx, row_data in enumerate(samples, 2):
            for col_idx, val in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

    @staticmethod
    def _get_headers(template_type: str):
        # User Spec Implementation
        if template_type == 'sales':
            return ['Date', 'Product_ID', 'Product_Name', 'Category', 'Sales_Qty', 'Price', 'Store', 'Promotion']
        elif template_type == 'inventory':
            return ['Date', 'SKU', 'Product_Name', 'Warehouse', 'Stock_Level', 'Reorder_Point', 'Lead_Time_Days', 'Cost_Per_Unit']
        elif template_type == 'financial':
            return ['Date', 'Account_Code', 'Account_Name', 'Department', 'Revenue', 'Expenses', 'Budget', 'Variance_Notes']
        return ['Date', 'Target_Value', 'Series_ID', 'Related_Metric']

    @staticmethod
    def _get_sample_data(template_type: str):
        # 10 realistic rows
        if template_type == 'sales':
            return [
                ['2024-01-01', 'P101', 'Widget A', 'Widgets', 150, 25.00, 'S001', 'No'],
                ['2024-01-02', 'P101', 'Widget A', 'Widgets', 142, 25.00, 'S001', 'No'],
                ['2024-01-03', 'P101', 'Widget A', 'Widgets', 160, 25.00, 'S001', 'Yes'],
                ['2024-01-04', 'P101', 'Widget A', 'Widgets', 155, 25.00, 'S001', 'No'],
                ['2024-01-05', 'P102', 'Gadget B', 'Gadgets', 89, 45.50, 'S002', 'Yes'],
                ['2024-01-06', 'P102', 'Gadget B', 'Gadgets', 92, 45.50, 'S002', 'No'],
                ['2024-01-07', 'P103', 'Gizmo C', 'Gizmos', 45, 120.00, 'S001', 'No'],
                ['2024-01-08', 'P103', 'Gizmo C', 'Gizmos', 47, 120.00, 'S001', 'No'],
                ['2024-01-09', 'P101', 'Widget A', 'Widgets', 130, 25.00, 'S003', 'No'],
                ['2024-01-10', 'P102', 'Gadget B', 'Gadgets', 100, 45.50, 'S003', 'Yes'],
            ]
        elif template_type == 'inventory':
             return [
                 ['2024-01-01', 'SH-001', 'Running Shoe', 'WH-East', 500, 100, 5, 25.00],
                 ['2024-01-02', 'SH-001', 'Running Shoe', 'WH-East', 480, 100, 5, 25.00],
                 ['2024-01-03', 'SH-001', 'Running Shoe', 'WH-East', 450, 100, 5, 25.00],
                 ['2024-01-01', 'SH-002', 'Hiking Boot', 'WH-West', 200, 50, 7, 45.00],
                 ['2024-01-02', 'SH-002', 'Hiking Boot', 'WH-West', 198, 50, 7, 45.00],
                 ['2024-01-03', 'SH-002', 'Hiking Boot', 'WH-West', 190, 50, 7, 45.00],
                 ['2024-01-01', 'GL-001', 'Gloves M', 'WH-Central', 1000, 200, 3, 10.00],
                 ['2024-01-02', 'GL-001', 'Gloves M', 'WH-Central', 950, 200, 3, 10.00],
                 ['2024-01-03', 'GL-001', 'Gloves M', 'WH-Central', 920, 200, 3, 10.00],
                 ['2024-01-04', 'GL-001', 'Gloves M', 'WH-Central', 890, 200, 3, 10.00],
             ]
        return []

    @staticmethod
    def _add_date_validation(ws, col_idx):
        col_letter = get_column_letter(col_idx)
        dv = DataValidation(type="date", operator="greaterThan", formula1="1900-01-01")
        dv.error = "Please enter a valid date (YYYY-MM-DD)"
        dv.errorTitle = "Invalid Date"
        ws.add_data_validation(dv)
        dv.add(f'{col_letter}2:{col_letter}10000')

    @staticmethod
    def _add_dropdown_validation(ws, col_idx, options):
        col_letter = get_column_letter(col_idx)
        options_str = ",".join(options)
        dv = DataValidation(type="list", formula1=f'"{options_str}"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f'{col_letter}2:{col_letter}10000')
